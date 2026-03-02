"""
Motor de exportación Excel — Informe Semanal GES-FO-016

Genera un archivo .xlsx que replica fielmente la plantilla oficial de Findeter
para informes semanales de interventoría.

Secciones:
  1. Información General (datos del contrato)
  2. Control de Hitos (snapshot inmutable)
  3. Situaciones Problemáticas
  4. Plan de Acción
  5. Actividades No Previstas (narrativa)
  6. Comentarios del Interventor
  7. Registro Fotográfico
"""

from __future__ import annotations

import os
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.contrato import (
    ActividadNoPrevista,
    ContratoInterventoria,
    ContratoObra,
    EstadoHito,
)
from app.models.informe import (
    AccionPlan,
    EstadoPlanAccion,
    InformeFoto,
    InformeSemanal,
    SnapshotHito,
)

# ---------------------------------------------------------------------------
# Constants — styling
# ---------------------------------------------------------------------------
ASSETS_DIR = Path(__file__).parent / "assets"

FONT_TITLE = Font(name="Arial", size=24, bold=True)
FONT_HEADER = Font(name="Arial", size=20, bold=True)
FONT_NORMAL = Font(name="Arial", size=20)
FONT_SECTION = Font(name="Arial", size=16, bold=True)
FONT_SECTION_BODY = Font(name="Arial", size=26)
FONT_SMALL = Font(name="Arial", size=16)
FONT_CODE = Font(name="Arial", size=20)

FILL_SECTION_S1S2 = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
FILL_SECTION_S3_S7 = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
FILL_GREEN = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

SIDE_THIN = Side(style="thin")
SIDE_MEDIUM = Side(style="medium")
SIDE_HAIR = Side(style="hair")
SIDE_DOTTED = Side(style="dotted")

BORDER_BOX = Border(left=SIDE_MEDIUM, right=SIDE_MEDIUM, top=SIDE_MEDIUM, bottom=SIDE_MEDIUM)
BORDER_THIN = Border(left=SIDE_THIN, right=SIDE_THIN, top=SIDE_THIN, bottom=SIDE_THIN)
BORDER_CELL = Border(left=SIDE_HAIR, right=SIDE_HAIR, top=SIDE_HAIR, bottom=SIDE_HAIR)
BORDER_DOTTED = Border(left=SIDE_DOTTED, right=SIDE_DOTTED, top=SIDE_DOTTED, bottom=SIDE_DOTTED)

# Column widths matching the template (A=1, S=19)
COL_WIDTHS = {
    "A": 21.0, "B": 20.85, "C": 70.71, "D": 9.28, "E": 15.71,
    "F": 26.57, "G": 24.85, "H": 15.71, "I": 22.0, "J": 15.71,
    "K": 26.0, "L": 10.28, "M": 16.85, "N": 10.28, "O": 3.14,
    "P": 19.71, "Q": 15.71, "R": 18.85, "S": 24.57,
}

DATE_FMT = "DD-MMM-YYYY"
MONEY_FMT = '#,##0'
PCT_FMT = '0.00%'


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _fmt_date(d: date | datetime | None) -> str:
    """Format a date to DD/MM/YYYY string."""
    if d is None:
        return ""
    if isinstance(d, datetime):
        d = d.date()
    return d.strftime("%d/%m/%Y")


def _fmt_money(v: Decimal | float | int | None) -> str:
    if v is None:
        return "$0"
    return f"${int(v):,}"


def _set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None,
              number_format=None):
    """Set a cell value with optional styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def _merge_and_set(ws, row_start, row_end, col_start, col_end, value,
                   font=None, fill=None, alignment=None, border=None, number_format=None):
    """Merge a range and set the top-left cell."""
    ws.merge_cells(
        start_row=row_start, start_column=col_start,
        end_row=row_end, end_column=col_end,
    )
    _set_cell(ws, row_start, col_start, value, font=font, fill=fill, alignment=alignment,
              border=border, number_format=number_format)
    # Apply border to all cells in the merged range
    if border:
        for r in range(row_start, row_end + 1):
            for c in range(col_start, col_end + 1):
                ws.cell(row=r, column=c).border = border


def _apply_border_range(ws, row_start, row_end, col_start, col_end, border):
    """Apply a border to all cells in a range."""
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).border = border


def _estado_hito_text(estado: EstadoHito) -> str:
    mapping = {
        EstadoHito.NO_INICIADO: "no iniciado",
        EstadoHito.EN_PROCESO: "EN PROCESO",
        EstadoHito.CUMPLIDO: "APROBADO",
        EstadoHito.VENCIDO: "VENCIDO",
    }
    return mapping.get(estado, str(estado.value))


def _estado_hito_fill(estado: EstadoHito) -> PatternFill | None:
    if estado == EstadoHito.CUMPLIDO:
        return FILL_GREEN
    elif estado == EstadoHito.VENCIDO:
        return FILL_RED
    elif estado == EstadoHito.EN_PROCESO:
        return FILL_YELLOW
    return None


# ---------------------------------------------------------------------------
# Section builders
# ---------------------------------------------------------------------------
def _build_header(ws, informe: InformeSemanal, contrato_obra: ContratoObra) -> int:
    """Build header with logo, title, and code. Returns the next available row."""
    # Logo (if available)
    logo_path = ASSETS_DIR / "logo_findeter.jpg"
    if logo_path.exists():
        try:
            img = XlImage(str(logo_path))
            img.width = 251
            img.height = 101
            ws.add_image(img, "A2")
        except Exception:
            pass  # Logo not available or corrupt — skip

    # Title
    ws.merge_cells("C4:K5")
    _set_cell(ws, 4, 3, "INFORME SEMANAL DE INTERVENTORÍA",
              font=FONT_TITLE, alignment=ALIGN_CENTER)

    # Code / Version / Date
    ws.merge_cells("P3:S3")
    _set_cell(ws, 3, 16, "Código: GES-FO-016", font=FONT_NORMAL, alignment=ALIGN_LEFT)
    ws.merge_cells("P4:S4")
    _set_cell(ws, 4, 16, "Versión: 3", font=FONT_NORMAL, alignment=ALIGN_LEFT)
    ws.merge_cells("P5:S5")
    _set_cell(ws, 5, 16, "Fecha de Aprobación: 14-Feb-2023",
              font=FONT_NORMAL, alignment=ALIGN_LEFT)
    ws.merge_cells("P6:S6")
    _set_cell(ws, 6, 16, "Clasificación: Pública", font=FONT_NORMAL, alignment=ALIGN_LEFT)

    # Row heights for header area
    ws.row_dimensions[6].height = 41.25
    ws.row_dimensions[8].height = 2.25
    ws.row_dimensions[10].height = 45.0
    ws.row_dimensions[11].height = 12.0

    return 12


def _build_s1(ws, informe: InformeSemanal, contrato_obra: ContratoObra,
              contrato_interv: ContratoInterventoria) -> int:
    """Section 1 — Información General. Returns next available row."""
    row = 12

    # Date / Period header row
    ws.row_dimensions[row].height = 35.1
    _merge_and_set(ws, row, row, 1, 2, "FECHA", font=FONT_HEADER, alignment=ALIGN_CENTER,
                   border=BORDER_THIN)
    _merge_and_set(ws, row, row, 3, 6, informe.semana_fin, font=FONT_NORMAL,
                   alignment=ALIGN_CENTER, border=BORDER_THIN)
    _merge_and_set(ws, row, row, 7, 8, "PERIODO No. ", font=FONT_HEADER,
                   alignment=ALIGN_CENTER, border=BORDER_THIN)
    _set_cell(ws, row, 9, informe.numero_informe, font=FONT_NORMAL, alignment=ALIGN_CENTER,
              border=BORDER_THIN)
    _set_cell(ws, row, 10, "DEL", font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
    _merge_and_set(ws, row, row, 11, 16, informe.semana_inicio, font=FONT_HEADER,
                   alignment=ALIGN_CENTER, border=BORDER_THIN)
    _set_cell(ws, row, 17, "AL", font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
    _merge_and_set(ws, row, row, 18, 19, informe.semana_fin, font=FONT_HEADER,
                   alignment=ALIGN_CENTER, border=BORDER_THIN)
    row += 1

    # "1. INFORMACIÓN GENERAL" header
    _merge_and_set(ws, row, row + 2, 1, 19, "1. INFORMACIÓN GENERAL ",
                   font=FONT_HEADER, fill=FILL_SECTION_S1S2, alignment=ALIGN_CENTER,
                   border=BORDER_THIN)
    ws.row_dimensions[row + 1].height = 42.0
    row += 3

    # OBJETO DEL CONTRATO (row 16)
    ws.row_dimensions[row].height = 114.0
    _merge_and_set(ws, row, row, 1, 3, "OBJETO DEL CONTRATO",
                   font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_CELL)
    _merge_and_set(ws, row, row, 4, 19, contrato_interv.objeto,
                   font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_CELL)
    row += 1

    # LOCALIZACIÓN (row 17)
    ws.row_dimensions[row].height = 39.95
    _merge_and_set(ws, row, row, 1, 3, "LOCALIZACIÓN DEL PROYECTO",
                   font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_CELL)
    _merge_and_set(ws, row, row, 4, 19, "LOCALIDAD DE BOSA – BOGOTÁ D.C.",
                   font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_CELL)
    row += 1

    # CONTRATO DE INTERVENTORÍA / CONTRATO DE OBRA headers (row 18)
    ws.row_dimensions[row].height = 60.95
    _merge_and_set(ws, row, row, 1, 7, "CONTRATO DE INTERVENTORÍA",
                   font=FONT_HEADER, alignment=ALIGN_CENTER, border=BORDER_THIN)
    _merge_and_set(ws, row, row, 8, 19, "CONTRATO DE OBRA",
                   font=FONT_HEADER, alignment=ALIGN_CENTER, border=BORDER_THIN)
    row += 1

    # Contract detail rows (19-32)
    detail_rows = [
        ("CONTRATO No.: ", contrato_interv.numero,
         "CONTRATO No.: ", contrato_obra.numero),
        ("PLAZO INICIAL: ", f"{contrato_interv.plazo_dias} DÍAS",
         "PLAZO INICIAL: ", f"{contrato_obra.plazo_dias} DÍAS"),
        ("FECHA DE INICIACIÓN:", contrato_interv.fecha_inicio,
         "FECHA DE INICIACIÓN:", contrato_obra.fecha_inicio),
        ("**FECHA DE SUSPENSIÓN:", "N.A." if not contrato_obra.fecha_suspension else contrato_obra.fecha_suspension,
         "**FECHA DE SUSPENSIÓN:", "N.A." if not contrato_obra.fecha_suspension else contrato_obra.fecha_suspension),
        ("**FECHA DE REINICIACIÓN:", "N.A." if not contrato_obra.fecha_reinicio else contrato_obra.fecha_reinicio,
         "**FECHA DE REINICIACIÓN:", "N.A." if not contrato_obra.fecha_reinicio else contrato_obra.fecha_reinicio),
        ("FECHA DE TERMINACIÓN:", contrato_interv.fecha_terminacion,
         "FECHA DE TERMINACIÓN:", contrato_obra.fecha_terminacion),
        ("PLAZO ACTUALIZADO: ", f"{contrato_interv.plazo_dias} DÍAS",
         "PLAZO ACTUALIZADO: ", f"{contrato_obra.plazo_dias} DÍAS"),
    ]

    for label_i, val_i, label_o, val_o in detail_rows:
        ws.row_dimensions[row].height = 30.0
        _merge_and_set(ws, row, row, 1, 3, label_i,
                       font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_CELL)
        # Format dates
        display_i = _fmt_date(val_i) if isinstance(val_i, (date, datetime)) else val_i
        display_o = _fmt_date(val_o) if isinstance(val_o, (date, datetime)) else val_o
        _merge_and_set(ws, row, row, 4, 7, display_i,
                       font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_CELL)
        _merge_and_set(ws, row, row, 8, 10, label_o,
                       font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_CELL)
        _merge_and_set(ws, row, row, 11, 19, display_o,
                       font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_CELL)
        row += 1

    # Plazo transcurrido
    ws.row_dimensions[row].height = 30.0
    dias_transcurridos_i = max(0, (informe.semana_fin - contrato_interv.fecha_inicio).days)
    dias_transcurridos_o = max(0, (informe.semana_fin - contrato_obra.fecha_inicio).days)
    _merge_and_set(ws, row, row, 1, 3, "PLAZO TRANSCURRIDO:",
                   font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_CELL)
    _merge_and_set(ws, row, row, 4, 5, dias_transcurridos_i,
                   font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_CELL)
    _merge_and_set(ws, row, row, 6, 7, "DÍAS",
                   font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_CELL)
    _merge_and_set(ws, row, row, 8, 10, "PLAZO TRANSCURRIDO:",
                   font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_CELL)
    _merge_and_set(ws, row, row, 11, 16, dias_transcurridos_o,
                   font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_CELL)
    _merge_and_set(ws, row, row, 17, 19, "DÍAS",
                   font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_CELL)
    row += 1

    # Financial rows
    financial_rows = [
        ("VALOR INICIAL :", contrato_interv.valor_inicial,
         "VALOR INICIAL :", contrato_obra.valor_inicial),
        ("VALOR ADICION(ES):", "N.A." if contrato_obra.adiciones == 0 else contrato_obra.adiciones,
         "VALOR ADICION(ES):", "N.A." if contrato_obra.adiciones == 0 else contrato_obra.adiciones),
        ("VALOR ACTUALIZADO:", contrato_interv.valor_actualizado,
         "VALOR ACTUALIZADO:", contrato_obra.valor_actualizado),
        ("VALOR PAGADO:", float(informe.valor_acumulado_ejecutado),
         "VALOR PAGADO:", float(informe.valor_acumulado_ejecutado)),
    ]

    for label_i, val_i, label_o, val_o in financial_rows:
        ws.row_dimensions[row].height = 30.0
        _merge_and_set(ws, row, row, 1, 3, label_i,
                       font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_CELL)
        if isinstance(val_i, str):
            display_i = val_i
        else:
            display_i = _fmt_money(val_i)
        if isinstance(val_o, str):
            display_o = val_o
        else:
            display_o = _fmt_money(val_o)
        _merge_and_set(ws, row, row, 4, 7, display_i,
                       font=FONT_NORMAL, alignment=ALIGN_RIGHT, border=BORDER_CELL)
        _merge_and_set(ws, row, row, 8, 10, label_o,
                       font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_CELL)
        _merge_and_set(ws, row, row, 11, 19, display_o,
                       font=FONT_NORMAL, alignment=ALIGN_RIGHT, border=BORDER_CELL)
        row += 1

    # VALOR POR PAGAR
    ws.row_dimensions[row].height = 30.0
    por_pagar_i = contrato_interv.valor_actualizado - informe.valor_acumulado_ejecutado
    por_pagar_o = contrato_obra.valor_actualizado - informe.valor_acumulado_ejecutado
    _merge_and_set(ws, row, row, 1, 3, "VALOR POR PAGAR:",
                   font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_CELL)
    _merge_and_set(ws, row, row, 4, 7, _fmt_money(por_pagar_i),
                   font=FONT_NORMAL, alignment=ALIGN_RIGHT, border=BORDER_CELL)
    _merge_and_set(ws, row, row, 8, 10, "VALOR POR PAGAR:",
                   font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_CELL)
    _merge_and_set(ws, row, row, 11, 19, _fmt_money(por_pagar_o),
                   font=FONT_NORMAL, alignment=ALIGN_RIGHT, border=BORDER_CELL)
    row += 1

    # SUPERVISOR / INTERVENTOR
    ws.row_dimensions[row].height = 30.0
    _merge_and_set(ws, row, row, 1, 3, "SUPERVISOR:",
                   font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_CELL)
    _merge_and_set(ws, row, row, 4, 7, contrato_interv.supervisor,
                   font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_CELL)
    _merge_and_set(ws, row, row, 8, 10, "INTERVENTOR",
                   font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_CELL)
    _merge_and_set(ws, row, row, 11, 19, contrato_interv.contratista,
                   font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_CELL)
    row += 1

    return row


def _build_s2(ws, informe: InformeSemanal, snapshots: list[SnapshotHito],
              start_row: int, contrato_obra: ContratoObra | None = None) -> int:
    """Section 2 — Control de Hitos. Returns next available row."""
    row = start_row + 1  # blank row separator

    # Section header
    _merge_and_set(ws, row, row + 1, 1, 19, "2. CONTROL DE HITOS.",
                   font=FONT_HEADER, fill=FILL_SECTION_S1S2, alignment=ALIGN_CENTER,
                   border=BORDER_THIN)
    ws.row_dimensions[row].height = 30.0
    row += 2

    # Table header
    ws.row_dimensions[row].height = 48.95
    _set_cell(ws, row, 1, "No.", font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_DOTTED)
    _merge_and_set(ws, row, row, 2, 9, "DESCRIPCIÓN DEL HITO",
                   font=FONT_HEADER, alignment=ALIGN_CENTER, border=BORDER_DOTTED)
    _merge_and_set(ws, row, row, 10, 11, "FECHA PROGRAMADA",
                   font=FONT_HEADER, alignment=ALIGN_CENTER, border=BORDER_DOTTED)
    # Columns L-O: estado/avance (merged implicitly with other cols in template)
    _set_cell(ws, row, 12, "ESTADO", font=FONT_HEADER, alignment=ALIGN_CENTER,
              border=BORDER_DOTTED)
    _set_cell(ws, row, 13, "AVANCE %", font=FONT_HEADER, alignment=ALIGN_CENTER,
              border=BORDER_DOTTED)
    _merge_and_set(ws, row, row, 16, 17, "FECHA REAL DE CUMPLIMIENTO",
                   font=FONT_HEADER, alignment=ALIGN_CENTER, border=BORDER_DOTTED)
    _merge_and_set(ws, row, row, 18, 19, "DÍAS DE RETRASO",
                   font=FONT_HEADER, alignment=ALIGN_CENTER, border=BORDER_DOTTED)
    row += 1

    # Hito rows
    for sh in sorted(snapshots, key=lambda h: h.numero):
        ws.row_dimensions[row].height = 35.1
        _set_cell(ws, row, 1, f"Hito {sh.numero}", font=FONT_NORMAL, alignment=ALIGN_CENTER,
                  border=BORDER_DOTTED)
        _merge_and_set(ws, row, row, 2, 9, sh.descripcion,
                       font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_DOTTED)
        _merge_and_set(ws, row, row, 10, 11,
                       _fmt_date(sh.fecha_programada),
                       font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_DOTTED)

        # Estado
        estado_text = _estado_hito_text(sh.estado)
        estado_fill = _estado_hito_fill(sh.estado)
        _set_cell(ws, row, 12, estado_text, font=FONT_NORMAL, alignment=ALIGN_CENTER,
                  border=BORDER_DOTTED, fill=estado_fill)

        # Avance
        avance_val = float(sh.avance_porcentaje) / 100.0 if sh.avance_porcentaje else 0
        _set_cell(ws, row, 13, avance_val, font=FONT_NORMAL, alignment=ALIGN_CENTER,
                  border=BORDER_DOTTED, number_format=PCT_FMT)

        # Fecha real
        _merge_and_set(ws, row, row, 16, 17,
                       _fmt_date(sh.fecha_real) if sh.fecha_real else "",
                       font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_DOTTED)

        # Días de retraso
        retraso_text = str(sh.dias_retraso) if sh.dias_retraso > 0 else (
            estado_text if sh.estado == EstadoHito.CUMPLIDO else "no iniciado"
            if sh.estado == EstadoHito.NO_INICIADO else ""
        )
        retraso_fill = None
        if sh.dias_retraso > 0:
            retraso_fill = FILL_RED
        elif sh.estado == EstadoHito.CUMPLIDO:
            retraso_fill = FILL_GREEN

        _merge_and_set(ws, row, row, 18, 19, retraso_text,
                       font=Font(name="Arial", size=20, bold=True),
                       alignment=ALIGN_CENTER, border=BORDER_DOTTED)
        if retraso_fill:
            for c in range(18, 20):
                ws.cell(row=row, column=c).fill = retraso_fill
        row += 1

    # Indicators section
    row += 1
    ws.row_dimensions[row].height = 30.0
    _set_cell(ws, row, 1, "No.", font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
    _merge_and_set(ws, row, row, 2, 8, "INDICADORES SEMANALES",
                   font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
    _merge_and_set(ws, row, row, 9, 10, "VALORES", font=FONT_NORMAL,
                   alignment=ALIGN_CENTER, border=BORDER_THIN)
    _merge_and_set(ws, row, row, 11, 16, "PORCENTAJE SOBRE VALOR TOTAL",
                   font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
    _merge_and_set(ws, row, row, 17, 19, "DIFERENCIA\n(+) ADELANTO\n(-) ATRASO",
                   font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
    row += 1

    valor_total = float(contrato_obra.valor_actualizado) if contrato_obra else 1

    # Row 1: Valor acumulado programado
    _set_cell(ws, row, 1, 1, font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
    _merge_and_set(ws, row, row, 2, 8,
                   "Valor acumulado de las actividades conforme a la programación vigente.",
                   font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
    val_prog = float(informe.valor_acumulado_programado)
    _merge_and_set(ws, row, row, 9, 10, _fmt_money(val_prog),
                   font=FONT_NORMAL, alignment=ALIGN_RIGHT, border=BORDER_THIN)
    pct_prog = val_prog / valor_total if valor_total else 0
    _merge_and_set(ws, row, row, 11, 16, pct_prog,
                   font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN,
                   number_format=PCT_FMT)
    row += 1

    # Row 2: Valor acumulado ejecutado
    _set_cell(ws, row, 1, 2, font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
    _merge_and_set(ws, row, row, 2, 8,
                   "Valor acumulado de las actividades ejecutadas y aprobadas por la Interventoría.",
                   font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
    val_ejec = float(informe.valor_acumulado_ejecutado)
    _merge_and_set(ws, row, row, 9, 10, _fmt_money(val_ejec),
                   font=FONT_NORMAL, alignment=ALIGN_RIGHT, border=BORDER_THIN)
    pct_ejec = val_ejec / valor_total if valor_total else 0
    _merge_and_set(ws, row, row, 11, 16, pct_ejec,
                   font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN,
                   number_format=PCT_FMT)
    # Diferencia financiera
    diff_fin = pct_ejec - pct_prog
    diff_font = Font(name="Arial", size=20, bold=True,
                     color="FF0000" if diff_fin < 0 else "008000")
    _merge_and_set(ws, row, row, 17, 19, diff_fin,
                   font=diff_font, alignment=ALIGN_CENTER, border=BORDER_THIN,
                   number_format=PCT_FMT)
    row += 1

    # Row 3: Avance físico programado
    _set_cell(ws, row, 1, 3, font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
    _merge_and_set(ws, row, row, 2, 8,
                   "Avance Físico Programado de las actividades conforme a la programación vigente.",
                   font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_THIN)
    avf_prog = float(informe.avance_fisico_programado) / 100.0
    _merge_and_set(ws, row, row, 9, 10, avf_prog,
                   font=Font(name="Arial", size=20, bold=True),
                   alignment=ALIGN_CENTER, border=BORDER_THIN, number_format=PCT_FMT)
    row += 1

    # Row 4: Avance físico ejecutado
    _set_cell(ws, row, 1, 4, font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
    _merge_and_set(ws, row, row, 2, 8,
                   "Avance Físico Ejecutado de las actividades conforme a la programación vigente.",
                   font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_THIN)
    avf_ejec = float(informe.avance_fisico_ejecutado) / 100.0
    _merge_and_set(ws, row, row, 9, 10, avf_ejec,
                   font=Font(name="Arial", size=20, bold=True),
                   alignment=ALIGN_CENTER, border=BORDER_THIN, number_format=PCT_FMT)
    # Diferencia física
    diff_fis = avf_ejec - avf_prog
    diff_fis_font = Font(name="Arial", size=20, bold=True,
                         color="FF0000" if diff_fis < 0 else "008000")
    _merge_and_set(ws, row, row, 17, 19, diff_fis,
                   font=diff_fis_font, alignment=ALIGN_CENTER, border=BORDER_THIN,
                   number_format=PCT_FMT)
    row += 1

    return row


def _build_s3(ws, informe: InformeSemanal, start_row: int) -> int:
    """Section 3 — Situaciones Problemáticas. Returns next available row."""
    row = start_row + 1

    # Header
    _merge_and_set(ws, row, row, 1, 19,
                   "3. IDENTIFICACIÓN DE SITUACIONES PROBLEMÁTICAS - ANÁLISIS DE CAUSAS",
                   font=FONT_SECTION, fill=FILL_SECTION_S3_S7, alignment=ALIGN_CENTER,
                   border=BORDER_THIN)
    ws.row_dimensions[row].height = 40.5
    row += 1

    # Content — split text into paragraphs
    text = informe.situaciones_problematicas or "(Sin situaciones problemáticas registradas)"
    paragraphs = [p.strip() for p in text.split("\n") if p.strip()]

    if not paragraphs:
        paragraphs = [text]

    for paragraph in paragraphs:
        _merge_and_set(ws, row, row, 1, 19, paragraph,
                       font=FONT_SECTION_BODY, alignment=ALIGN_LEFT_TOP,
                       border=BORDER_THIN)
        # Estimate row height based on text length (rough)
        estimated_chars_per_line = 120
        lines = max(1, len(paragraph) // estimated_chars_per_line + paragraph.count("\n") + 1)
        ws.row_dimensions[row].height = max(40, lines * 30)
        row += 1

    return row


def _build_s4(ws, informe: InformeSemanal, acciones: list[AccionPlan],
              start_row: int) -> int:
    """Section 4 — Plan de Acción. Returns next available row."""
    row = start_row

    # Header
    _merge_and_set(ws, row, row, 1, 19,
                   "4. PLAN DE ACCIÓN RESULTADO DEL ANÁLISIS DE CAUSAS",
                   font=FONT_SECTION, fill=FILL_SECTION_S3_S7, alignment=ALIGN_CENTER,
                   border=BORDER_THIN)
    ws.row_dimensions[row].height = 69.0
    row += 1

    # Column headers
    ws.row_dimensions[row].height = 80.25
    _merge_and_set(ws, row, row, 1, 8, "ACTIVIDAD",
                   font=FONT_SECTION, alignment=ALIGN_CENTER, border=BORDER_THIN)
    _merge_and_set(ws, row, row, 9, 15, "RESPONSABLE",
                   font=FONT_SECTION, alignment=ALIGN_CENTER, border=BORDER_THIN)
    _merge_and_set(ws, row, row, 16, 19, "FECHA\nPROGRAMADA",
                   font=FONT_SECTION, alignment=ALIGN_CENTER, border=BORDER_THIN)
    row += 1

    # Action rows
    for accion in acciones:
        ws.row_dimensions[row].height = 66.75
        _merge_and_set(ws, row, row, 1, 8, accion.actividad,
                       font=FONT_NORMAL, alignment=ALIGN_LEFT, border=BORDER_THIN)
        _merge_and_set(ws, row, row, 9, 15, accion.responsable,
                       font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
        fecha_text = _fmt_date(accion.fecha_programada) if accion.fecha_programada else "Permanente"
        _merge_and_set(ws, row, row, 16, 19, fecha_text,
                       font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
        row += 1

    if not acciones:
        _merge_and_set(ws, row, row, 1, 19, "(Sin acciones registradas)",
                       font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
        ws.row_dimensions[row].height = 40
        row += 1

    return row


def _build_s5(ws, informe: InformeSemanal, actividades_np: list[ActividadNoPrevista],
              start_row: int) -> int:
    """Section 5 — Actividades No Previstas. Returns next available row."""
    row = start_row

    # Header
    _merge_and_set(ws, row, row, 1, 19,
                   "5. ACTIVIDADES NO PREVISTAS Y MAYORES CANTIDADES",
                   font=FONT_SECTION, fill=FILL_SECTION_S3_S7, alignment=ALIGN_CENTER,
                   border=BORDER_THIN)
    ws.row_dimensions[row].height = 79.5
    row += 1

    # Narrative text
    narrativa = informe.actividades_no_previstas_narrativa or informe.actividades_no_previstas or ""
    paragraphs = [p.strip() for p in narrativa.split("\n") if p.strip()]

    if not paragraphs:
        paragraphs = ["(Sin actividades no previstas registradas)"]

    for paragraph in paragraphs:
        _merge_and_set(ws, row, row, 1, 19, paragraph,
                       font=FONT_SECTION_BODY, alignment=ALIGN_LEFT_TOP,
                       border=BORDER_THIN)
        estimated_lines = max(1, len(paragraph) // 120 + 1)
        ws.row_dimensions[row].height = max(40, estimated_lines * 30)
        row += 1

    return row


def _build_s6(ws, informe: InformeSemanal, start_row: int) -> int:
    """Section 6 — Comentarios del Interventor. Returns next available row."""
    row = start_row

    # Header
    _merge_and_set(ws, row, row, 1, 19, "6. COMENTARIOS DEL INTERVENTOR",
                   font=FONT_SECTION, fill=FILL_SECTION_S3_S7, alignment=ALIGN_CENTER,
                   border=BORDER_THIN)
    ws.row_dimensions[row].height = 51.0
    row += 1

    # Combine all comments into paragraphs
    comments = []
    if informe.comentario_tecnico:
        comments.append(informe.comentario_tecnico)
    if informe.comentario_sst:
        comments.append(f"SST: {informe.comentario_sst}")
    if informe.comentario_ambiental:
        comments.append(f"AMBIENTAL: {informe.comentario_ambiental}")
    if informe.comentario_social:
        comments.append(f"SOCIAL: {informe.comentario_social}")

    if not comments:
        comments = ["(Sin comentarios registrados)"]

    for comment in comments:
        paragraphs = [p.strip() for p in comment.split("\n") if p.strip()]
        for paragraph in paragraphs:
            _merge_and_set(ws, row, row, 1, 19, paragraph,
                           font=FONT_SECTION_BODY, alignment=ALIGN_LEFT_TOP,
                           border=BORDER_THIN)
            estimated_lines = max(1, len(paragraph) // 120 + 1)
            ws.row_dimensions[row].height = max(40, estimated_lines * 30)
            row += 1

    return row


def _build_s7(ws, informe: InformeSemanal, fotos_seleccionadas: list[InformeFoto],
              upload_dir: str, start_row: int) -> int:
    """Section 7 — Registro Fotográfico. Returns next available row."""
    row = start_row

    # Header
    _merge_and_set(ws, row, row, 1, 19,
                   "7. REGISTRO FOTOGRÁFICO DEL AVANCE DURANTE LA SEMANA ",
                   font=FONT_SECTION, fill=FILL_SECTION_S3_S7, alignment=ALIGN_CENTER,
                   border=BORDER_THIN)
    ws.row_dimensions[row].height = 21.95
    row += 1

    if not fotos_seleccionadas:
        _merge_and_set(ws, row, row, 1, 19, "(Sin registro fotográfico)",
                       font=FONT_NORMAL, alignment=ALIGN_CENTER, border=BORDER_THIN)
        ws.row_dimensions[row].height = 40
        row += 1
        return row

    # Photos in 3-column grid
    # Each photo block: 15 rows for image + 1 row for caption
    PHOTO_ROWS = 15
    COLS_GRID = [
        (1, 3),    # columns A-C
        (5, 9),    # columns E-I
        (11, 19),  # columns K-S
    ]

    fotos_sorted = sorted(fotos_seleccionadas, key=lambda f: f.orden)

    i = 0
    while i < len(fotos_sorted):
        # Process up to 3 photos per row
        batch = fotos_sorted[i:i + 3]

        # Image rows
        img_start = row
        img_end = row + PHOTO_ROWS - 1

        for idx, foto_sel in enumerate(batch):
            col_start, col_end = COLS_GRID[idx]

            # Merge the image area
            ws.merge_cells(
                start_row=img_start, start_column=col_start,
                end_row=img_end, end_column=col_end,
            )

            # Set row heights for image area
            for r in range(img_start, img_end + 1):
                ws.row_dimensions[r].height = 21.95

            # Try to insert actual image
            foto = foto_sel.foto
            img_path = os.path.join(upload_dir, foto.archivo_path)
            if os.path.exists(img_path):
                try:
                    img = XlImage(img_path)
                    # Scale to fit ~300x250 px area
                    max_w, max_h = 500, 350
                    if img.width > max_w or img.height > max_h:
                        ratio = min(max_w / img.width, max_h / img.height)
                        img.width = int(img.width * ratio)
                        img.height = int(img.height * ratio)
                    cell_ref = f"{get_column_letter(col_start)}{img_start}"
                    ws.add_image(img, cell_ref)
                except Exception:
                    # If image can't be loaded, put placeholder text
                    ws.cell(row=img_start, column=col_start).value = f"[Foto: {foto.archivo_nombre}]"
            else:
                ws.cell(row=img_start, column=col_start).value = f"[Foto: {foto.archivo_nombre}]"
                ws.cell(row=img_start, column=col_start).font = FONT_SMALL
                ws.cell(row=img_start, column=col_start).alignment = ALIGN_CENTER

        row = img_end + 1

        # Caption row
        for idx, foto_sel in enumerate(batch):
            col_start, col_end = COLS_GRID[idx]
            caption = foto_sel.pie_de_foto_override or foto_sel.foto.pie_de_foto or foto_sel.foto.archivo_nombre
            _merge_and_set(ws, row, row, col_start, col_end, caption,
                           font=FONT_SECTION, alignment=ALIGN_CENTER, border=BORDER_THIN)
        ws.row_dimensions[row].height = 40.5
        row += 1

        # Small gap
        ws.row_dimensions[row].height = 21.95
        row += 1

        i += len(batch)

    return row


def _build_signatures(ws, contrato_interv: ContratoInterventoria, start_row: int) -> int:
    """Build signature section. Returns next available row."""
    row = start_row + 2  # gap

    ws.row_dimensions[row].height = 74.25
    row += 1

    # Director signature
    _set_cell(ws, row, 1, "FIRMA DIRECTOR DE INTERVENTORÍA",
              font=FONT_SECTION, alignment=ALIGN_LEFT)
    ws.row_dimensions[row].height = 30.75
    row += 1

    _merge_and_set(ws, row, row, 1, 19, f"NOMBRE: {contrato_interv.contratista.upper()}",
                   font=FONT_SECTION, alignment=ALIGN_LEFT)
    ws.row_dimensions[row].height = 25.5
    row += 2

    # Supervisor signature
    ws.row_dimensions[row].height = 72.75
    row += 1
    ws.row_dimensions[row].height = 63.0
    row += 1
    ws.row_dimensions[row].height = 81.0
    row += 1

    _merge_and_set(ws, row, row, 1, 19,
                   f"NOMBRE SUPERVISOR: {contrato_interv.supervisor.upper()}",
                   font=FONT_SECTION, alignment=ALIGN_LEFT)
    ws.row_dimensions[row].height = 44.25
    row += 1

    return row


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------
def generate_informe_excel(
    db: Session,
    informe: InformeSemanal,
    upload_dir: str | None = None,
) -> BytesIO:
    """
    Generate an Excel file (.xlsx) for the given InformeSemanal.

    Returns a BytesIO buffer with the workbook content.
    """
    if upload_dir is None:
        from app.core.config import settings
        upload_dir = settings.UPLOAD_DIR

    # Load related data
    contrato_obra: ContratoObra = informe.contrato_obra
    contrato_interv: ContratoInterventoria = contrato_obra.contrato_interventoria

    snapshots: list[SnapshotHito] = informe.snapshot_hitos
    fotos_sel: list[InformeFoto] = informe.fotos_seleccionadas

    # Load acciones plan (this report + pending from previous)
    acciones = (
        db.query(AccionPlan)
        .filter(
            AccionPlan.informe_origen_id == informe.id,
        )
        .order_by(AccionPlan.numero)
        .all()
    )

    # Also get pending actions from previous reports of the same contrato
    from sqlalchemy import and_, or_

    acciones_previas = (
        db.query(AccionPlan)
        .join(InformeSemanal, AccionPlan.informe_origen_id == InformeSemanal.id)
        .filter(
            and_(
                InformeSemanal.contrato_obra_id == informe.contrato_obra_id,
                InformeSemanal.id != informe.id,
                InformeSemanal.semana_inicio < informe.semana_inicio,
                or_(
                    AccionPlan.estado == EstadoPlanAccion.PENDIENTE,
                    AccionPlan.estado == EstadoPlanAccion.EN_PROCESO,
                ),
            )
        )
        .order_by(AccionPlan.numero)
        .all()
    )
    all_acciones = acciones_previas + acciones

    # Load actividades no previstas
    actividades_np = (
        db.query(ActividadNoPrevista)
        .filter(ActividadNoPrevista.contrato_obra_id == informe.contrato_obra_id)
        .order_by(ActividadNoPrevista.codigo)
        .all()
    )

    # --- Build workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = f"Inf No {informe.numero_informe} Cto {contrato_obra.numero}"

    # Set column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Page setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Build sections
    row = _build_header(ws, informe, contrato_obra)
    row = _build_s1(ws, informe, contrato_obra, contrato_interv)
    row = _build_s2(ws, informe, snapshots, row, contrato_obra=contrato_obra)
    row = _build_s3(ws, informe, row)
    row = _build_s4(ws, informe, all_acciones, row)
    row = _build_s5(ws, informe, actividades_np, row)
    row = _build_s6(ws, informe, row)
    row = _build_s7(ws, informe, fotos_sel, upload_dir, row)
    _build_signatures(ws, contrato_interv, row)

    # Write to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
